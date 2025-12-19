import time
import hmac
import hashlib
import requests
import json
import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# --- Load secrets ---
load_dotenv()
API_KEY = os.getenv("API_KEY")
API_SECRET = os.getenv("API_SECRET")

# --- Settings ---
BASE_URL = "https://api.coindcx.com"
PAIR = "SOLINR"  # changed to INR pair
PRICE_THRESHOLD = 13300.0  # adjust your trigger price here in INR
BUY_AMOUNT_INR = 500.0    # amount to spend in INR per buy order
PROFIT_TARGET_PERCENT = 10
CHECK_INTERVAL = 30
EXCEL_LOG_FILE = "sol_inr_trade_log.xlsx"

# --- Logging Function ---
def log(action, message, buy_price=None, sell_price=None, quantity=None, profit=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(EXCEL_LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade Logs"
        ws.append(["Timestamp", "Action", "Buy Price", "Sell Price", "Quantity", "Profit (INR)", "Message"])
        wb.save(EXCEL_LOG_FILE)

    wb = load_workbook(EXCEL_LOG_FILE)
    ws = wb.active
    ws.append([
        timestamp,
        action,
        round(buy_price, 2) if buy_price else "",
        round(sell_price, 2) if sell_price else "",
        round(quantity, 3) if quantity else "",
        round(profit, 2) if profit else "",
        message
    ])
    wb.save(EXCEL_LOG_FILE)

    print(f"[{timestamp}] [{action}] {message}")

# --- Market Price Fetcher ---
def get_market_price():
    url = "https://api.coindcx.com/exchange/ticker"
    response = requests.get(url)
    tickers = response.json()
    for t in tickers:
        if t["market"] == PAIR:
            return float(t["last_price"])
    return None

# --- Authentication Header Generator ---
def get_headers(payload):
    timestamp = int(time.time() * 1000)
    signature = hmac.new(
        API_SECRET.encode(),
        payload.encode(),
        hashlib.sha256
    ).hexdigest()
    return {
        'X-AUTH-APIKEY': API_KEY,
        'X-AUTH-SIGNATURE': signature,
        'X-AUTH-TIMESTAMP': str(timestamp),
        'Content-Type': 'application/json'
    }

# --- Place Limit Order ---
def place_limit_order(side, price, quantity):
    url = BASE_URL + "/exchange/v1/orders/create"
    body = {
        "market": PAIR,
        "side": side,
        "order_type": "limit_order",
        "price_per_unit": str(price),
        "total_quantity": format(quantity, '.3f'),
        "timestamp": int(time.time() * 1000)
    }
    payload = json.dumps(body)
    headers = get_headers(payload)
    response = requests.post(url, headers=headers, data=payload)

    if response.status_code == 200:
        log("ORDER", f"{side.upper()} order placed at ₹{price}",
            buy_price=price if side == "buy" else None,
            sell_price=price if side == "sell" else None,
            quantity=quantity)
        return True
    else:
        log("ERROR", f"Order failed: {response.text}")
        return False

# --- Main Bot Loop ---
def main():
    while True:
        try:
            current_price = get_market_price()
            log("INFO", f"Current SOL price: ₹{current_price:.2f}")

            if current_price and current_price < PRICE_THRESHOLD:
                buy_price = round(current_price - 5.0, 2)
                quantity = float(format(BUY_AMOUNT_INR / buy_price, '.3f'))

                if place_limit_order("buy", buy_price, quantity):
                    sell_price = round(buy_price * (1 + PROFIT_TARGET_PERCENT / 100), 2)
                    profit = (sell_price - buy_price) * quantity
                    place_limit_order("sell", sell_price, quantity)
                    log("ORDER", "Sell order placed", buy_price=buy_price, sell_price=sell_price,
                        quantity=quantity, profit=profit)
                    time.sleep(60)  # avoid duplicate orders

            else:
                log("INFO", "Price above threshold, waiting...")

        except Exception as e:
            log("ERROR", str(e))

        time.sleep(CHECK_INTERVAL)

# --- Start the Bot ---
if __name__ == "__main__":
    main()
